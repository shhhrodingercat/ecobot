# Questo software è rilasciato sotto la Licenza Pubblica Generica GNU v3 (GPL v3)
# Per maggiori informazioni, vedere il file LICENSE o https://www.gnu.org/licenses/gpl-3.0.html

import os
import fitz
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Impostazioni delle variabili
pdf_folder_path = ""  # Inserisci il percorso della cartella con i file PDF
y_base = 100
incremento_y = 15.35
larghezza_rettangolo = 90
altezza_rettangolo = 16
distanza_a_destra = 10

# Definizione dei colori e delle relative etichette
colori_bollini = {
    (144, 91, 54): "Organico",
    (254, 206, 63): "Carta",
    (0, 125, 187): "Vetro",
    (26, 185, 236): "Plastica",
    (124, 126, 128): "Indifferenziato",
    (108, 189, 71): "Sfalci"
}

# Tolleranza dei colori
tolleranza_colori = {name: 10 for name in colori_bollini.values()}
colonne = ["Indifferenziato", "Organico", "Carta", "Vetro", "Plastica", "Sfalci"]

def calcola_distanza_colore(colore1, colore2):
    """Calcola la distanza euclidea tra due colori RGB"""
    return sum((a - b) ** 2 for a, b in zip(colore1, colore2)) ** 0.5

def analizza_colori(immagine, colori_riferimento, tolleranze):
    """
    Analizza i colori presenti nell'immagine e restituisce una lista dei colori trovati
    """
    colori_trovati = set()
    larghezza, altezza = immagine.size

    # Converti l'immagine in RGB se non lo è già
    if immagine.mode != 'RGB':
        immagine = immagine.convert('RGB')

    # Analizza i pixel dell'immagine
    for x in range(larghezza):
        for y in range(altezza):
            pixel_colore = immagine.getpixel((x, y))

            # Confronta con i colori di riferimento
            for colore_rif, nome_colore in colori_riferimento.items():
                distanza = calcola_distanza_colore(pixel_colore, colore_rif)
                if distanza <= tolleranze[nome_colore]:
                    colori_trovati.add(nome_colore)

    return list(colori_trovati)

# Crea un nuovo libro di lavoro Excel
wb = Workbook()

def analizza_pdf(pdf_path, sheet_name):
    pdf_document = fitz.open(pdf_path)
    ws = wb.create_sheet(title=sheet_name)  # Crea un nuovo foglio con il nome del file

    # Definizione intestazione
    intestazione = ["date", "i", "o", "c", "v", "p", "s"]
    ws.append(intestazione)

    # Definizione coordinate per le due pagine
    prima_pagina = [(25, 29), (165, 31), (300, 30), (440, 31), (580, 30), (715, 31)]
    seconda_pagina = [(25, 31), (165, 30), (300, 31), (440, 30), (580, 31), (715, 31)]

    def analizza_mese(pagina_corrente, data_mese, mese_coords):
        y_start = y_base - incremento_y
        i = 1
        for mese in mese_coords:
            x_mese, giorni_del_mese = mese
            for giorno in range(giorni_del_mese):
                y_mese = y_start + (incremento_y * i)
                x_analisi = x_mese + distanza_a_destra
                y_analisi = y_mese - altezza_rettangolo / 2

                # Ottieni l'immagine della pagina
                immagine_pagina = pagina_corrente.get_pixmap()
                immagine_pillow = Image.frombytes("RGB",
                                                [immagine_pagina.width, immagine_pagina.height],
                                                immagine_pagina.samples)

                # Ritaglia il rettangolo di interesse
                rettangolo_immagine = immagine_pillow.crop((
                    x_analisi, y_analisi,
                    x_analisi + larghezza_rettangolo,
                    y_analisi + altezza_rettangolo
                ))

                # Genera la data (modificato per 2025-2026)
                data = f"2025-{data_mese:02d}-{i:02d}" if data_mese < 13 else f"2026-01-{i:02d}"

                # Analizza i colori presenti
                colori_presenti = analizza_colori(rettangolo_immagine, colori_bollini, tolleranza_colori)
                colori_presenti_bool = [int(colonna in colori_presenti) for colonna in colonne]

                # Aggiungi la riga al foglio Excel
                ws.append([data] + colori_presenti_bool)
                print(f"{data}: {colori_presenti}")
                i += 1
            data_mese += 1

    # Analizza entrambe le pagine
    analizza_mese(pdf_document[0], 2, prima_pagina)  # Inizia da febbraio 2025
    analizza_mese(pdf_document[1], 8, seconda_pagina)  # Seconda pagina da agosto 2025

    pdf_document.close()

# Elabora tutti i file PDF nella cartella
for file_name in os.listdir(pdf_folder_path):
    if file_name.endswith(".pdf"):
        pdf_path = os.path.join(pdf_folder_path, file_name)
        sheet_name = os.path.splitext(file_name)[0]  # Usa il nome del file come nome del foglio
        try:
            analizza_pdf(pdf_path, sheet_name)
            print(f"Elaborato con successo: {file_name}")
        except Exception as e:
            print(f"Errore nell'elaborazione di {file_name}: {str(e)}")

# Rimuovi il foglio di default se esiste
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# Salva il file Excel
output_path = "database.xlsx"  # Modificato il nome del file di output
wb.save(output_path)
print(f"File Excel salvato in: {output_path}")
